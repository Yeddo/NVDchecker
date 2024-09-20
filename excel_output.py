import openpyxl # Excel libraries
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def save_to_excel(cve_results, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CVE Report" # Names the Worksheet
    ws.merge_cells('A1:F2') # Merges top row A1 - F2
    cell = ws.cell(row=1, column=1) # Sets which cell, cell is.
    cell.value = "CVE Report" # Puts this in cell
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True, size=18)
    # This moves the row down because for some reason the merge doesn't get seen as 2 rows, only 1 but the next appended row ends up blank if this isn't here. ¯\_(ツ)_/¯
    ws.append([""])
    # This needs to be here before changing font style otherwise it'll "append" below first row. Or you can specify where every cell goes that the column headers go in individually (ws['A1'] = "Package Version").
    headers = ["Package Name", "Package Version", "CVE ID", "CVSS Score", "Published Date", "Description"]
    ws.append(headers)
    # Bold header row.
    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    ws['C3'].font = Font(bold=True)
    ws['D3'].font = Font(bold=True)
    ws['E3'].font = Font(bold=True)
    ws['F3'].font = Font(bold=True)
    # Center header row except for last row 'F'
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['B3'].alignment = Alignment(horizontal='center')
    ws['C3'].alignment = Alignment(horizontal='center')
    ws['D3'].alignment = Alignment(horizontal='center')
    ws['E3'].alignment = Alignment(horizontal='center')
    # Resize all the columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

#    for i, header in enumerate(headers, 1):
#            ws.column_dimensions[get_column_letter(i)].width = 20

    for result in cve_results:
        ws.append(result)

    wb.save(output_file)
